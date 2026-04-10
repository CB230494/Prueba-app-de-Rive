# app.py
# -*- coding: utf-8 -*-

import re
import math
import unicodedata
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image as RLImage,
    PageBreak,
    KeepTogether,
)
from reportlab.graphics.shapes import Drawing, Rect, String

# =========================================================
# CONFIGURACIÓN GENERAL
# =========================================================
st.set_page_config(
    page_title="Seguimiento de líneas de acción",
    layout="wide"
)

# =========================================================
# SESSION STATE
# =========================================================
if "lineas_guardadas" not in st.session_state:
    st.session_state["lineas_guardadas"] = {}

if "archivo_nombre" not in st.session_state:
    st.session_state["archivo_nombre"] = ""

if "pdf_final" not in st.session_state:
    st.session_state["pdf_final"] = None

if "excel_final" not in st.session_state:
    st.session_state["excel_final"] = None

if "ultimo_resumen" not in st.session_state:
    st.session_state["ultimo_resumen"] = pd.DataFrame()

# =========================================================
# ESTILOS CSS
# =========================================================
st.markdown("""
<style>
.block-container {
    max-width: 1550px;
    padding-top: 1rem;
    padding-bottom: 2rem;
}

.main-title {
    font-size: 2rem;
    font-weight: 700;
    margin-bottom: 0.2rem;
    color: inherit;
}

.sub-note {
    font-size: 0.95rem;
    color: rgba(250,250,250,0.78);
    margin-bottom: 1rem;
}

.info-card {
    border: 1px solid rgba(255,255,255,0.08);
    background: rgba(255,255,255,0.03);
    border-radius: 14px;
    padding: 16px 18px;
    margin-bottom: 16px;
}

.line-card {
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 14px;
    padding: 18px;
    margin-bottom: 22px;
    background: rgba(255,255,255,0.02);
    box-shadow: 0 1px 2px rgba(0,0,0,0.18);
}

.section-title {
    font-size: 1.12rem;
    font-weight: 700;
    margin-bottom: 0.6rem;
    color: inherit;
}

.kpi-box {
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 12px;
    padding: 12px;
    background: rgba(255,255,255,0.03);
    text-align: center;
    color: inherit;
    min-height: 82px;
    display: flex;
    flex-direction: column;
    justify-content: center;
}

.small-note {
    font-size: 0.9rem;
    color: rgba(250,250,250,0.72);
}

hr.soft-line {
    border: none;
    height: 1px;
    background: rgba(255,255,255,0.12);
    margin-top: 10px;
    margin-bottom: 14px;
}
</style>
""", unsafe_allow_html=True)

# =========================================================
# UTILIDADES DE TEXTO
# =========================================================
def strip_accents(text: str) -> str:
    if text is None:
        return ""
    text = str(text)
    return "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )


def normalize_text(value) -> str:
    if value is None:
        return ""

    s = str(value)
    s = strip_accents(s).lower()
    s = s.replace("\n", " ").replace("\r", " ")
    s = s.replace("°", " ")
    s = s.replace("№", " ")
    s = re.sub(r"[|]+", " ", s)
    s = re.sub(r"[_]+", " ", s)
    s = re.sub(r"[-]+", " ", s)
    s = re.sub(r"[#]+", " # ", s)
    s = re.sub(r"[:;,]+", " ", s)
    s = re.sub(r"[()]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_nonempty(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def compact_join(values):
    parts = []
    seen = set()
    for v in values:
        t = clean_text(v)
        n = normalize_text(t)
        if t and n and n not in seen:
            parts.append(t)
            seen.add(n)
    return " / ".join(parts)


# =========================================================
# PALABRAS CLAVE
# =========================================================
KEY_LINEA = [
    "linea de accion",
    "linea accion",
    "linea de coordinacion",
    "linea de accion #",
    "linea accion #",
]

KEY_PROBLEMATICA = [
    "problematica",
    "problematica de linea de accion",
    "problematica de linea",
    "problematica de la linea",
]

KEY_LIDER = [
    "lider estrategico",
    "lider",
]

KEY_HEADER_INDICADOR = ["indicador", "indicadores"]
KEY_HEADER_META = ["meta"]
KEY_HEADER_AVANCE = ["avance"]
KEY_HEADER_DESCRIPCION = ["descripcion", "descripcio"]
KEY_HEADER_CANTIDAD = ["cantidad"]
KEY_HEADER_OBSERVACIONES = ["observaciones", "observacion", "obs"]

BAD_LINE_TEXT = [
    "problematica",
    "lider",
    "delegacion",
    "trimestre",
    "indicador",
    "meta",
    "avance",
    "descripcion",
    "cantidad",
    "observacion",
    "linea de accion",
]

# =========================================================
# UTILIDADES DE EXCEL
# =========================================================
def get_effective_cell_value(ws, row, col):
    if row < 1 or col < 1:
        return None

    cell = ws.cell(row=row, column=col)

    if not isinstance(cell, MergedCell):
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return ws.cell(merged_range.min_row, merged_range.min_col).value

    return None


def row_values(ws, row, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    return [get_effective_cell_value(ws, row, c) for c in range(1, max_col + 1)]


def row_text(ws, row, max_col=None):
    vals = row_values(ws, row, max_col)
    return " | ".join("" if v is None else str(v) for v in vals)


def get_right_nonempty(ws, row, col, max_steps=12):
    for c in range(col + 1, min(ws.max_column, col + max_steps) + 1):
        val = get_effective_cell_value(ws, row, c)
        if is_nonempty(val):
            return val
    return ""


def get_left_nonempty(ws, row, col, max_steps=8):
    for c in range(col - 1, max(1, col - max_steps), -1):
        val = get_effective_cell_value(ws, row, c)
        if is_nonempty(val):
            return val
    return ""


def get_down_nonempty(ws, row, col, max_steps=6):
    for r in range(row + 1, min(ws.max_row, row + max_steps) + 1):
        val = get_effective_cell_value(ws, r, col)
        if is_nonempty(val):
            return val
    return ""


def get_up_nonempty(ws, row, col, max_steps=6):
    for r in range(row - 1, max(1, row - max_steps), -1):
        val = get_effective_cell_value(ws, r, col)
        if is_nonempty(val):
            return val
    return ""


def get_near_nonempty(ws, row, col):
    for func, steps in [
        (get_right_nonempty, 12),
        (get_down_nonempty, 6),
        (get_left_nonempty, 8),
        (get_up_nonempty, 6),
    ]:
        val = func(ws, row, col, max_steps=steps)
        if is_nonempty(val):
            return val
    return ""


def sheet_density_score(ws, max_row=220, max_col=40):
    score = 0
    mr = min(ws.max_row, max_row)
    mc = min(ws.max_column, max_col)

    for r in range(1, mr + 1):
        txt = normalize_text(row_text(ws, r, mc))
        score += 8 if "linea de accion" in txt else 0
        score += 6 if "problematica" in txt else 0
        score += 6 if "lider estrategico" in txt or "lider" in txt else 0
        score += 4 if "indicador" in txt else 0
        score += 4 if "meta" in txt else 0
        score += 4 if "avance" in txt else 0
        score += 4 if "descripcion" in txt else 0
        score += 4 if "cantidad" in txt else 0
        score += 4 if "observaciones" in txt else 0
        score += 3 if "delegacion" in txt else 0
        score += 2 if "trimestre" in txt else 0

    return score


# =========================================================
# DETECCIÓN DE HOJA
# =========================================================
def find_best_main_sheet(wb):
    candidates = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        score = sheet_density_score(ws)

        name_norm = normalize_text(sheet_name)

        if "informe de avance" in name_norm:
            score += 50
        if "informe" in name_norm:
            score += 20
        if "avance" in name_norm:
            score += 20
        if "dashboard" in name_norm or "datos" in name_norm or "sumatoria" in name_norm:
            score -= 20

        candidates.append((sheet_name, score))

    candidates.sort(key=lambda x: x[1], reverse=True)
    return candidates[0][0] if candidates else wb.sheetnames[0]


# =========================================================
# DATOS GENERALES
# =========================================================
def get_delegacion(ws):
    max_row = min(ws.max_row, 80)
    max_col = min(ws.max_column, 25)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = get_effective_cell_value(ws, r, c)
            txt = normalize_text(val)

            if "delegacion" in txt:
                near = get_near_nonempty(ws, r, c)
                if is_nonempty(near):
                    cleaned = clean_text(near)
                    if normalize_text(cleaned) != "delegacion":
                        return cleaned

    return ""


def get_fecha_actualizacion(ws):
    max_row = min(ws.max_row, 30)
    max_col = min(ws.max_column, 25)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = get_effective_cell_value(ws, r, c)
            txt = normalize_text(val)

            if "fecha de actualizacion" in txt:
                near = get_near_nonempty(ws, r, c)
                if is_nonempty(near):
                    return clean_text(near)

    return ""


# =========================================================
# LÍNEAS
# =========================================================
def looks_like_bad_line_value(text: str) -> bool:
    t = normalize_text(text)
    if not t:
        return True
    return any(p in t for p in BAD_LINE_TEXT)


def extract_line_number_from_area(ws, start_row, start_col):
    own_text = clean_text(get_effective_cell_value(ws, start_row, start_col))
    own_norm = normalize_text(own_text)

    m = re.search(r"linea\s+de\s+accion\s*#?\s*(\d+)", own_norm)
    if m:
        return m.group(1)

    candidates = []

    for c in range(start_col, min(ws.max_column, start_col + 10) + 1):
        val = get_effective_cell_value(ws, start_row, c)
        if is_nonempty(val):
            txt = clean_text(val)
            if not looks_like_bad_line_value(txt):
                candidates.append(txt)

    for r in range(start_row, min(ws.max_row, start_row + 3) + 1):
        for c in range(max(1, start_col - 1), min(ws.max_column, start_col + 8) + 1):
            val = get_effective_cell_value(ws, r, c)
            if is_nonempty(val):
                txt = clean_text(val)
                if not looks_like_bad_line_value(txt):
                    candidates.append(txt)

    for candidate in candidates:
        m = re.search(r"(\d+(?:[.\-]\d+)?)", str(candidate))
        if m:
            return m.group(1)

    return ""


def find_line_action_starts(ws):
    starts = []
    mr = ws.max_row
    mc = min(ws.max_column, 25)

    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            val = get_effective_cell_value(ws, r, c)
            txt = normalize_text(val)

            if any(k in txt for k in KEY_LINEA):
                line_value = extract_line_number_from_area(ws, r, c)
                starts.append({
                    "row": r,
                    "col": c,
                    "line_value": line_value
                })
                break

    cleaned = []
    last_row = -999

    for item in starts:
        if item["row"] - last_row > 2:
            cleaned.append(item)
            last_row = item["row"]

    return cleaned


def search_value_near_keywords_multiline(ws, start_row, end_row, keywords, value_blacklist=None):
    if value_blacklist is None:
        value_blacklist = []

    for r in range(start_row, min(end_row, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            val = get_effective_cell_value(ws, r, c)
            txt = normalize_text(val)

            if any(k in txt for k in keywords):
                primary = get_right_nonempty(ws, r, c, 12)
                if not is_nonempty(primary):
                    primary = get_down_nonempty(ws, r, c, 3)

                base = clean_text(primary)
                collected = []

                if base:
                    collected.append(base)

                for rr in range(r + 1, min(ws.max_row, r + 2) + 1):
                    for cc in range(max(1, c), min(ws.max_column, c + 5) + 1):
                        extra = get_effective_cell_value(ws, rr, cc)
                        extra_clean = clean_text(extra)
                        extra_norm = normalize_text(extra_clean)

                        if not extra_norm:
                            continue
                        if any(b in extra_norm for b in value_blacklist):
                            continue
                        if "linea de accion" in extra_norm:
                            continue
                        if "trimestre" in extra_norm:
                            continue
                        if "indicador" in extra_norm:
                            continue
                        if len(extra_clean) < 4:
                            continue

                        if extra_norm != normalize_text(base):
                            collected.append(extra_clean)

                result = compact_join(collected)
                return result

    return ""


def detect_trimester(ws, start_row, end_row):
    roman_map = {
        "i": "I",
        "ii": "II",
        "iii": "III",
        "iv": "IV",
        "1": "I",
        "2": "II",
        "3": "III",
        "4": "IV",
        "primer": "I",
        "segundo": "II",
        "tercer": "III",
        "cuarto": "IV",
    }

    for r in range(start_row, min(end_row, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            val = get_effective_cell_value(ws, r, c)
            txt = normalize_text(val)

            if "trimestre" in txt:
                candidates = [
                    get_right_nonempty(ws, r, c, 5),
                    get_down_nonempty(ws, r, c, 2),
                    get_left_nonempty(ws, r, c, 2),
                ]

                for cand in candidates:
                    ct = normalize_text(cand)
                    for k, out in roman_map.items():
                        if re.fullmatch(rf"{re.escape(k)}", ct) or f"{k} trimestre" in ct:
                            return out

                row_txt = normalize_text(row_text(ws, r))
                if "iv trimestre" in row_txt or "4 trimestre" in row_txt or "cuarto trimestre" in row_txt:
                    return "IV"
                if "iii trimestre" in row_txt or "3 trimestre" in row_txt or "tercer trimestre" in row_txt:
                    return "III"
                if "ii trimestre" in row_txt or "2 trimestre" in row_txt or "segundo trimestre" in row_txt:
                    return "II"
                if "i trimestre" in row_txt or "1 trimestre" in row_txt or "primer trimestre" in row_txt:
                    return "I"

    return ""


# =========================================================
# TABLA
# =========================================================
def detect_header_row(ws, start_row, end_row):
    best_row = None
    best_score = -1

    for r in range(start_row, min(end_row, ws.max_row) + 1):
        vals = row_values(ws, r)

        found = {
            "indicador": False,
            "meta": False,
            "avance": False,
            "descripcion": False,
            "cantidad": False,
            "observaciones": False,
        }

        for v in vals:
            t = normalize_text(v)

            if any(k in t for k in KEY_HEADER_INDICADOR):
                found["indicador"] = True
            if any(k in t for k in KEY_HEADER_META):
                found["meta"] = True
            if any(k in t for k in KEY_HEADER_AVANCE):
                found["avance"] = True
            if any(k in t for k in KEY_HEADER_DESCRIPCION):
                found["descripcion"] = True
            if any(k in t for k in KEY_HEADER_CANTIDAD):
                found["cantidad"] = True
            if any(k in t for k in KEY_HEADER_OBSERVACIONES):
                found["observaciones"] = True

        score = sum(found.values())

        if found["indicador"] and found["meta"] and found["avance"] and found["descripcion"]:
            score += 4

        if score > best_score:
            best_score = score
            best_row = r

    if best_score < 3:
        return None

    return best_row


def map_headers(ws, header_row):
    header_map = {}

    for c in range(1, ws.max_column + 1):
        val = get_effective_cell_value(ws, header_row, c)
        t = normalize_text(val)

        if any(k in t for k in KEY_HEADER_INDICADOR) and "Indicador" not in header_map:
            header_map["Indicador"] = c

        elif any(k in t for k in KEY_HEADER_META) and "Meta (editable)" not in header_map:
            header_map["Meta (editable)"] = c

        elif any(k in t for k in KEY_HEADER_AVANCE) and "Avance (Editable)" not in header_map:
            header_map["Avance (Editable)"] = c

        elif any(k in t for k in KEY_HEADER_DESCRIPCION) and "Descripción (editable)" not in header_map:
            header_map["Descripción (editable)"] = c

        elif any(k in t for k in KEY_HEADER_CANTIDAD) and "Cantidad (editable)" not in header_map:
            header_map["Cantidad (editable)"] = c

        elif any(k in t for k in KEY_HEADER_OBSERVACIONES) and "Observaciones (Editable)" not in header_map:
            header_map["Observaciones (Editable)"] = c

    return header_map


def normalize_status_value(value):
    t = normalize_text(value)
    if not t:
        return ""

    if "completo" in t:
        return "Completado"
    if "con actividades" in t:
        return "Con Actividades"
    if "sin actividades" in t:
        return "Sin Actividades"
    return clean_text(value)


def extract_table(ws, header_row, block_end_row):
    columns = [
        "Indicador",
        "Meta (editable)",
        "Avance (Editable)",
        "Descripción (editable)",
        "Cantidad (editable)",
        "Observaciones (Editable)"
    ]

    header_map = map_headers(ws, header_row)

    if "Indicador" not in header_map:
        return pd.DataFrame([{
            "Indicador": "",
            "Meta (editable)": "",
            "Avance (Editable)": "",
            "Descripción (editable)": "",
            "Cantidad (editable)": "",
            "Observaciones (Editable)": ""
        }])

    data = []
    empty_count = 0

    for r in range(header_row + 1, block_end_row + 1):
        row_data = {}

        indicador = get_effective_cell_value(ws, r, header_map.get("Indicador")) if header_map.get("Indicador") else ""
        meta = get_effective_cell_value(ws, r, header_map.get("Meta (editable)")) if header_map.get("Meta (editable)") else ""
        avance = get_effective_cell_value(ws, r, header_map.get("Avance (Editable)")) if header_map.get("Avance (Editable)") else ""
        descripcion = get_effective_cell_value(ws, r, header_map.get("Descripción (editable)")) if header_map.get("Descripción (editable)") else ""
        cantidad = get_effective_cell_value(ws, r, header_map.get("Cantidad (editable)")) if header_map.get("Cantidad (editable)") else ""
        observaciones = get_effective_cell_value(ws, r, header_map.get("Observaciones (Editable)")) if header_map.get("Observaciones (Editable)") else ""

        row_data["Indicador"] = "" if indicador is None else str(indicador)
        row_data["Meta (editable)"] = "" if meta is None else str(meta)
        row_data["Avance (Editable)"] = "" if avance is None else normalize_status_value(avance)
        row_data["Descripción (editable)"] = "" if descripcion is None else str(descripcion)
        row_data["Cantidad (editable)"] = "" if cantidad is None else str(cantidad)
        row_data["Observaciones (Editable)"] = "" if observaciones is None else str(observaciones)

        has_content = any(str(v).strip() != "" for v in row_data.values())

        if not has_content:
            empty_count += 1
            if empty_count >= 5:
                break
            continue

        if normalize_text(row_data["Indicador"]) in ["indicador", "indicadores"]:
            continue

        empty_count = 0
        data.append(row_data)

    if not data:
        return pd.DataFrame([{
            "Indicador": "",
            "Meta (editable)": "",
            "Avance (Editable)": "",
            "Descripción (editable)": "",
            "Cantidad (editable)": "",
            "Observaciones (Editable)": ""
        }])

    return pd.DataFrame(data, columns=columns)


def prepare_editor_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy()

    expected_columns = [
        "Indicador",
        "Meta (editable)",
        "Avance (Editable)",
        "Descripción (editable)",
        "Cantidad (editable)",
        "Observaciones (Editable)"
    ]

    for col in expected_columns:
        if col not in df2.columns:
            df2[col] = ""

    for col in expected_columns:
        df2[col] = df2[col].fillna("").astype(str)

    return df2[expected_columns]


def dataframe_has_real_content(df: pd.DataFrame) -> bool:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return False

    for _, row in df.iterrows():
        if any(str(v).strip() != "" for v in row.values):
            return True
    return False


# =========================================================
# EXTRACCIÓN COMPLETA
# =========================================================
def extract_blocks_from_sheet(ws):
    starts = find_line_action_starts(ws)
    delegacion = get_delegacion(ws)
    fecha_actualizacion = get_fecha_actualizacion(ws)

    blocks = []

    if not starts:
        return {
            "delegacion": delegacion,
            "fecha_actualizacion": fecha_actualizacion,
            "blocks": []
        }

    for i, start in enumerate(starts):
        start_row = start["row"]
        end_row = starts[i + 1]["row"] - 1 if i + 1 < len(starts) else ws.max_row

        linea_numero = start["line_value"] if start["line_value"] else str(i + 1)

        problematica = search_value_near_keywords_multiline(
            ws=ws,
            start_row=start_row,
            end_row=min(start_row + 15, end_row),
            keywords=KEY_PROBLEMATICA,
            value_blacklist=["problematica", "lider", "trimestre", "indicador", "meta"]
        )

        lider = search_value_near_keywords_multiline(
            ws=ws,
            start_row=start_row,
            end_row=min(start_row + 12, end_row),
            keywords=KEY_LIDER,
            value_blacklist=["lider", "trimestre", "indicador", "meta", "problematica"]
        )

        trimestre = detect_trimester(
            ws=ws,
            start_row=start_row,
            end_row=min(start_row + 12, end_row)
        )

        header_row = detect_header_row(
            ws=ws,
            start_row=start_row,
            end_row=min(start_row + 35, end_row)
        )

        tabla = extract_table(ws, header_row, end_row) if header_row else pd.DataFrame([{
            "Indicador": "",
            "Meta (editable)": "",
            "Avance (Editable)": "",
            "Descripción (editable)": "",
            "Cantidad (editable)": "",
            "Observaciones (Editable)": ""
        }])

        tabla = prepare_editor_dataframe(tabla)

        problematica_final = clean_text(problematica)
        lider_final = clean_text(lider)

        blocks.append({
            "delegacion": delegacion,
            "fecha_actualizacion": fecha_actualizacion,
            "linea_accion": str(linea_numero).strip(),
            "problematica": problematica_final if problematica_final else "Sin nombre en Excel",
            "lider": lider_final if lider_final else "",
            "trimestre": clean_text(trimestre),
            "tabla": tabla,
            "rango_inicio": start_row,
            "rango_fin": end_row,
            "header_row": header_row,
            "nombre_vacio_en_origen": not bool(problematica_final)
        })

    # filtrar líneas fantasma
    blocks_filtrados = []
    for b in blocks:
        if dataframe_has_real_content(b["tabla"]):
            blocks_filtrados.append(b)

    return {
        "delegacion": delegacion,
        "fecha_actualizacion": fecha_actualizacion,
        "blocks": blocks_filtrados
    }


# =========================================================
# MÉTRICAS / SEMÁFORO
# =========================================================
def compute_line_metrics(df: pd.DataFrame) -> dict:
    total = 0
    completos = 0
    con_actividades = 0
    sin_actividades = 0

    if isinstance(df, pd.DataFrame):
        total = len(df.index)

        if "Avance (Editable)" in df.columns:
            for v in df["Avance (Editable)"].fillna("").astype(str):
                n = normalize_status_value(v)
                if n == "Completado":
                    completos += 1
                elif n == "Con Actividades":
                    con_actividades += 1
                elif n == "Sin Actividades":
                    sin_actividades += 1

    porcentaje = 0.0
    if total > 0:
        porcentaje = ((completos + 0.5 * con_actividades) / total) * 100

    if porcentaje >= 80:
        estado = "Alto"
        color = "#1B5E20"   # verde
    elif porcentaje >= 50:
        estado = "Medio"
        color = "#EF6C00"   # naranja
    else:
        estado = "Bajo"
        color = "#B71C1C"   # rojo

    return {
        "total": total,
        "completos": completos,
        "con_actividades": con_actividades,
        "sin_actividades": sin_actividades,
        "porcentaje": round(porcentaje, 1),
        "estado": estado,
        "color": color,
    }


def build_summary_dataframe(data_lineas):
    rows = []

    for _, item in data_lineas.items():
        df = item["tabla"].copy()
        m = compute_line_metrics(df)

        rows.append({
            "Línea": item.get("display_linea", ""),
            "Problemática": item["info"].get("problematica", ""),
            "Líder Estratégico": item["info"].get("lider", ""),
            "Trimestre": item.get("trimestre", ""),
            "Indicadores": m["total"],
            "Completado": m["completos"],
            "Con Actividades": m["con_actividades"],
            "Sin Actividades": m["sin_actividades"],
            "% Avance": m["porcentaje"],
            "Semáforo": m["estado"],
        })

    return pd.DataFrame(rows)


# =========================================================
# EXCEL CONSOLIDADO
# =========================================================
def build_excel_export(data_lineas, delegacion_general, fecha_actualizacion=""):
    output = BytesIO()

    resumen = build_summary_dataframe(data_lineas)

    detalle_rows = []
    for _, item in data_lineas.items():
        linea = item.get("display_linea", "")
        problematica = item["info"].get("problematica", "")
        lider = item["info"].get("lider", "")
        trimestre = item.get("trimestre", "")
        df = prepare_editor_dataframe(item["tabla"])

        for i, (_, row) in enumerate(df.iterrows(), start=1):
            detalle_rows.append({
                "Delegación": delegacion_general,
                "Fecha de actualización": fecha_actualizacion,
                "Línea": linea,
                "Item": f"{linea}.{i}",
                "Problemática": problematica,
                "Líder Estratégico": lider,
                "Trimestre": trimestre,
                "Indicador": row.get("Indicador", ""),
                "Meta": row.get("Meta (editable)", ""),
                "Avance": row.get("Avance (Editable)", ""),
                "Descripción": row.get("Descripción (editable)", ""),
                "Cantidad": row.get("Cantidad (editable)", ""),
                "Observaciones": row.get("Observaciones (Editable)", ""),
            })

    detalle_df = pd.DataFrame(detalle_rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        resumen.to_excel(writer, index=False, sheet_name="Resumen")
        detalle_df.to_excel(writer, index=False, sheet_name="Detalle")

        meta_df = pd.DataFrame([{
            "Delegación": delegacion_general,
            "Fecha de actualización": fecha_actualizacion,
            "Líneas guardadas": len(data_lineas),
        }])
        meta_df.to_excel(writer, index=False, sheet_name="Datos generales")

    output.seek(0)
    return output.getvalue()


# =========================================================
# PDF PROFESIONAL
# =========================================================
def draw_progress_bar(percent: float, width=230, height=12, fill_color="#1B5E20"):
    percent = max(0, min(100, percent))
    drawing = Drawing(width, height + 14)
    drawing.add(Rect(0, 8, width, height, strokeColor=colors.HexColor("#B0BEC5"), fillColor=colors.HexColor("#ECEFF1")))
    drawing.add(Rect(0, 8, width * (percent / 100.0), height, strokeColor=None, fillColor=colors.HexColor(fill_color)))
    drawing.add(String(width + 8, 7, f"{percent:.1f}%", fontSize=8))
    return drawing


def draw_semaforo(label: str, color_hex: str):
    drawing = Drawing(90, 18)
    drawing.add(Rect(0, 2, 70, 12, fillColor=colors.HexColor(color_hex), strokeColor=colors.HexColor(color_hex)))
    drawing.add(String(75, 4, label, fontSize=8))
    return drawing


def get_logo_path():
    possible = [
        Path("001.png"),
        Path("./001.png"),
        Path("/mount/src/001.png"),
    ]
    for p in possible:
        if p.exists():
            return str(p)
    return None


def build_pdf_all_lines(data_lineas, delegacion_general, fecha_actualizacion=""):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=34,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "title_main",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=19,
        leading=23,
        textColor=colors.HexColor("#1B5E20"),
        spaceAfter=8
    )

    title_style2 = ParagraphStyle(
        "title_sub",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#EF6C00"),
        spaceAfter=10
    )

    normal_style = ParagraphStyle(
        "normal_custom",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        alignment=TA_LEFT,
        spaceAfter=4
    )

    small_style = ParagraphStyle(
        "small_custom",
        parent=styles["Normal"],
        fontSize=7.6,
        leading=9
    )

    heading_style = ParagraphStyle(
        "heading_custom",
        parent=styles["Heading2"],
        fontSize=12,
        leading=15,
        textColor=colors.HexColor("#B71C1C"),
        spaceAfter=8
    )

    elements = []

    # =========================================
    # PORTADA
    # =========================================
    logo_path = get_logo_path()
    elements.append(Spacer(1, 35))

    if logo_path:
        try:
            img = RLImage(logo_path, width=4.2 * inch, height=2.4 * inch)
            img.hAlign = "CENTER"
            elements.append(img)
            elements.append(Spacer(1, 25))
        except Exception:
            pass

    elements.append(Paragraph("REPORTE TRIMESTRAL", title_style))
    elements.append(Paragraph("SEGUIMIENTO DE LÍNEAS DE ACCIÓN", title_style))
    elements.append(Paragraph("Estrategia de Coordinación Estratégica", title_style2))
    elements.append(Spacer(1, 18))

    portada_table = Table([
        ["Delegación", safe_str(delegacion_general) or "-"],
        ["Fecha de actualización", safe_str(fecha_actualizacion) or "-"],
        ["Líneas incluidas", str(len(data_lineas))],
    ], colWidths=[160, 300])

    portada_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#1B5E20")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
        ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#F5F5F5")),
        ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#9E9E9E")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    elements.append(Spacer(1, 18))
    elements.append(portada_table)
    elements.append(Spacer(1, 40))

    intro = (
        "Este informe presenta el seguimiento trimestral de las líneas de acción registradas, "
        "con una estructura que permite visualizar de forma clara el estado de avance, la "
        "descripción operativa, la cantidad reportada y las observaciones asociadas a cada indicador. "
        "La información se organiza por línea de acción, incorporando un resumen ejecutivo, "
        "semáforo de avance y detalle consolidado para facilitar la revisión y la toma de decisiones."
    )
    elements.append(Paragraph(intro, normal_style))
    elements.append(PageBreak())

    # =========================================
    # RESUMEN EJECUTIVO
    # =========================================
    elements.append(Paragraph("RESUMEN EJECUTIVO", heading_style))
    resumen_df = build_summary_dataframe(data_lineas)

    resumen_table_data = [[
        "Línea", "Problemática", "Indicadores", "% Avance", "Semáforo"
    ]]

    for _, row in resumen_df.iterrows():
        resumen_table_data.append([
            Paragraph(safe_str(row.get("Línea", "")), small_style),
            Paragraph(safe_str(row.get("Problemática", "")), small_style),
            Paragraph(safe_str(row.get("Indicadores", "")), small_style),
            Paragraph(safe_str(row.get("% Avance", "")), small_style),
            Paragraph(safe_str(row.get("Semáforo", "")), small_style),
        ])

    resumen_table = Table(
        resumen_table_data,
        repeatRows=1,
        colWidths=[50, 250, 70, 70, 80]
    )
    resumen_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5E20")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 16))

    # =========================================
    # DETALLE POR LÍNEA
    # =========================================
    for key, item in data_lineas.items():
        info = item["info"]
        df = prepare_editor_dataframe(item["tabla"])
        trimestre = item.get("trimestre", "")
        display_linea = item.get("display_linea", key)
        metrics = compute_line_metrics(df)

        elements.append(Paragraph(f"LÍNEA DE ACCIÓN {safe_str(display_linea)}", heading_style))
        elements.append(Paragraph(f"<b>Problemática:</b> {safe_str(info.get('problematica', ''))}", normal_style))
        elements.append(Paragraph(f"<b>Líder Estratégico:</b> {safe_str(info.get('lider', ''))}", normal_style))
        elements.append(Paragraph(f"<b>Trimestre:</b> {safe_str(trimestre)}", normal_style))
        elements.append(Spacer(1, 6))

        metric_table = Table([
            ["Indicadores", str(metrics["total"]), "Completado", str(metrics["completos"])],
            ["Con Actividades", str(metrics["con_actividades"]), "Sin Actividades", str(metrics["sin_actividades"])],
        ], colWidths=[120, 60, 120, 60])

        metric_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F5F5")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EF6C00")),
            ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#1B5E20")),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
            ("TEXTCOLOR", (2, 0), (2, -1), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        elements.append(metric_table)
        elements.append(Spacer(1, 8))

        progress_and_semaforo = Table([
            [draw_progress_bar(metrics["porcentaje"], fill_color=metrics["color"]), draw_semaforo(metrics["estado"], metrics["color"])]
        ], colWidths=[300, 180])

        progress_and_semaforo.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        elements.append(progress_and_semaforo)
        elements.append(Spacer(1, 10))

        table_data = [[
            "Ítem", "Indicador", "Meta", "Avance", "Descripción", "Cantidad", "Observaciones"
        ]]

        for i, (_, row) in enumerate(df.iterrows(), start=1):
            item_num = f"{display_linea}.{i}"
            table_data.append([
                Paragraph(item_num, small_style),
                Paragraph(safe_str(row.get("Indicador", "")), small_style),
                Paragraph(safe_str(row.get("Meta (editable)", "")), small_style),
                Paragraph(safe_str(row.get("Avance (Editable)", "")), small_style),
                Paragraph(safe_str(row.get("Descripción (editable)", "")), small_style),
                Paragraph(safe_str(row.get("Cantidad (editable)", "")), small_style),
                Paragraph(safe_str(row.get("Observaciones (Editable)", "")), small_style),
            ])

        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[34, 92, 55, 62, 120, 50, 117]
        )

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5E20")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.1),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
        ]))

        elements.append(KeepTogether([table]))
        elements.append(Spacer(1, 14))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


# =========================================================
# INTERFAZ
# =========================================================
st.markdown('<div class="main-title">Seguimiento de líneas de acción</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="sub-note">Sube un archivo Excel (.xlsm o .xlsx). La app localizará de forma robusta la delegación, línea de acción, problemática, líder estratégico, trimestre y el detalle editable de indicadores.</div>',
    unsafe_allow_html=True
)

uploaded_file = st.file_uploader(
    "Arrastra y suelta el archivo Excel",
    type=["xlsm", "xlsx"]
)

if uploaded_file is not None:
    try:
        if st.session_state["archivo_nombre"] != uploaded_file.name:
            st.session_state["archivo_nombre"] = uploaded_file.name
            st.session_state["lineas_guardadas"] = {}
            st.session_state["pdf_final"] = None
            st.session_state["excel_final"] = None
            st.session_state["ultimo_resumen"] = pd.DataFrame()

        file_bytes = uploaded_file.read()

        wb = load_workbook(
            BytesIO(file_bytes),
            data_only=False,
            keep_vba=True
        )

        main_sheet = find_best_main_sheet(wb)
        ws = wb[main_sheet]

        extraction = extract_blocks_from_sheet(ws)

        delegacion = extraction["delegacion"]
        fecha_actualizacion = extraction["fecha_actualizacion"]
        blocks = extraction["blocks"]

        if not blocks:
            st.warning("No se encontraron bloques reales de líneas de acción en la hoja cargada.")
            st.stop()

        col_k1, col_k2, col_k3, col_k4 = st.columns(4)

        with col_k1:
            st.markdown(
                f'<div class="kpi-box"><b>Hoja detectada</b><br>{safe_str(main_sheet)}</div>',
                unsafe_allow_html=True
            )
        with col_k2:
            st.markdown(
                f'<div class="kpi-box"><b>Líneas detectadas</b><br>{len(blocks)}</div>',
                unsafe_allow_html=True
            )
        with col_k3:
            st.markdown(
                f'<div class="kpi-box"><b>Delegación</b><br>{safe_str(delegacion) or "-"}</div>',
                unsafe_allow_html=True
            )
        with col_k4:
            st.markdown(
                f'<div class="kpi-box"><b>Fecha actualización</b><br>{safe_str(fecha_actualizacion) or "-"}</div>',
                unsafe_allow_html=True
            )

        st.markdown("")

        st.markdown('<div class="info-card">', unsafe_allow_html=True)
        g1, g2 = st.columns([1.2, 4])

        with g1:
            st.text_input(
                "Delegación",
                value=delegacion,
                disabled=True,
                key="delegacion_general"
            )

        with g2:
            st.text_input(
                "Fecha de actualización",
                value=fecha_actualizacion,
                disabled=True,
                key="fecha_actualizacion_general"
            )

        st.markdown("</div>", unsafe_allow_html=True)

        with st.sidebar:
            st.header("Resumen")
            st.write(f"**Archivo:** {uploaded_file.name}")
            st.write(f"**Hoja detectada:** {main_sheet}")
            st.write(f"**Delegación:** {delegacion or '-'}")
            st.write(f"**Fecha de actualización:** {fecha_actualizacion or '-'}")
            st.write(f"**Líneas encontradas:** {len(blocks)}")
            st.divider()

            st.write("**Líneas detectadas:**")
            for b in blocks:
                nombre = b["problematica"]
                if b.get("nombre_vacio_en_origen"):
                    st.write(f"• Línea {b['linea_accion']} — Sin nombre en Excel")
                else:
                    st.write(f"• Línea {b['linea_accion']} — {nombre}")

        for idx, bloque in enumerate(blocks):
            linea_id = str(bloque["linea_accion"]).strip() if bloque["linea_accion"] else str(idx + 1)

            block_key = f"bloque_{idx}_{linea_id}"
            save_key = f"{idx}_{linea_id}"

            if save_key in st.session_state["lineas_guardadas"]:
                df_base = st.session_state["lineas_guardadas"][save_key]["tabla"].copy()
                trim_base = st.session_state["lineas_guardadas"][save_key]["trimestre"]
            else:
                df_base = bloque["tabla"].copy()
                trim_base = bloque["trimestre"] if bloque["trimestre"] in ["", "I", "II", "III", "IV"] else ""

            df_base = prepare_editor_dataframe(df_base)

            st.markdown('<div class="line-card">', unsafe_allow_html=True)
            st.markdown(f'<div class="section-title">Línea {safe_str(linea_id)}</div>', unsafe_allow_html=True)
            st.markdown('<hr class="soft-line">', unsafe_allow_html=True)

            c1, c2, c3 = st.columns([1.2, 3.8, 2])

            with c1:
                st.text_input(
                    "Línea de acción #",
                    value=linea_id,
                    disabled=True,
                    key=f"linea_{block_key}"
                )

            with c2:
                st.text_input(
                    "Problemática",
                    value=bloque["problematica"],
                    disabled=True,
                    key=f"problematica_{block_key}"
                )

            with c3:
                st.text_input(
                    "Líder Estratégico",
                    value=bloque["lider"],
                    disabled=True,
                    key=f"lider_{block_key}"
                )

            if bloque.get("nombre_vacio_en_origen"):
                st.warning(f"La línea {linea_id} viene sin nombre en el Excel original.")

            c4, c5 = st.columns([1.2, 5])

            with c4:
                trim_options = ["", "I", "II", "III", "IV"]
                selected_trim = st.selectbox(
                    "Trimestre",
                    trim_options,
                    index=trim_options.index(trim_base if trim_base in trim_options else ""),
                    key=f"trim_{block_key}"
                )

            with c5:
                st.caption(
                    f"Filas detectadas del bloque: {bloque['rango_inicio']} a {bloque['rango_fin']} | "
                    f"Encabezado detectado: {bloque['header_row'] if bloque['header_row'] else '-'}"
                )

            m = compute_line_metrics(df_base)
            s1, s2, s3, s4, s5 = st.columns(5)
            s1.metric("Indicadores", m["total"])
            s2.metric("Completado", m["completos"])
            s3.metric("Con Actividades", m["con_actividades"])
            s4.metric("Sin Actividades", m["sin_actividades"])
            s5.metric("% Avance", f"{m['porcentaje']}%")

            st.markdown("#### Detalle editable")

            df_editado = st.data_editor(
                df_base,
                use_container_width=True,
                hide_index=True,
                num_rows="dynamic",
                key=f"tabla_{block_key}",
                column_config={
                    "Indicador": st.column_config.TextColumn("Indicador", width="medium"),
                    "Meta (editable)": st.column_config.TextColumn("Meta (editable)", width="medium"),
                    "Avance (Editable)": st.column_config.SelectboxColumn(
                        "Avance (Editable)",
                        options=["", "Completado", "Con Actividades", "Sin Actividades"],
                        width="small"
                    ),
                    "Descripción (editable)": st.column_config.TextColumn("Descripción (editable)", width="large"),
                    "Cantidad (editable)": st.column_config.TextColumn("Cantidad (editable)", width="small"),
                    "Observaciones (Editable)": st.column_config.TextColumn("Observaciones (Editable)", width="large"),
                }
            )

            btn1, btn2 = st.columns([2.3, 5])

            with btn1:
                if st.button(f"Guardar / Actualizar línea {linea_id}", key=f"guardar_{block_key}"):
                    st.session_state["lineas_guardadas"][save_key] = {
                        "display_linea": linea_id,
                        "info": {
                            "delegacion": delegacion,
                            "linea_accion": linea_id,
                            "problematica": bloque["problematica"],
                            "lider": bloque["lider"],
                            "rango_inicio": bloque["rango_inicio"],
                            "rango_fin": bloque["rango_fin"],
                        },
                        "tabla": prepare_editor_dataframe(df_editado),
                        "trimestre": selected_trim
                    }
                    st.session_state["pdf_final"] = None
                    st.session_state["excel_final"] = None
                    st.success(f"Línea {linea_id} guardada correctamente.")

            with btn2:
                if st.button(f"Restaurar línea {linea_id}", key=f"restaurar_{block_key}"):
                    if save_key in st.session_state["lineas_guardadas"]:
                        del st.session_state["lineas_guardadas"][save_key]
                    st.session_state["pdf_final"] = None
                    st.session_state["excel_final"] = None
                    st.warning(f"Línea {linea_id} restaurada a los datos detectados del Excel.")
                    st.rerun()

            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("## Gestión general")

        gbtn1, gbtn2 = st.columns([2.5, 5])

        with gbtn1:
            if st.button("Guardar todas las líneas detectadas"):
                for idx, bloque in enumerate(blocks):
                    linea_id = str(bloque["linea_accion"]).strip() if bloque["linea_accion"] else str(idx + 1)
                    block_key = f"bloque_{idx}_{linea_id}"
                    save_key = f"{idx}_{linea_id}"

                    trim_value = st.session_state.get(f"trim_{block_key}", bloque["trimestre"])
                    tabla_value = st.session_state.get(f"tabla_{block_key}", bloque["tabla"])

                    if isinstance(tabla_value, pd.DataFrame):
                        tabla_guardar = prepare_editor_dataframe(tabla_value)
                    else:
                        tabla_guardar = prepare_editor_dataframe(bloque["tabla"])

                    st.session_state["lineas_guardadas"][save_key] = {
                        "display_linea": linea_id,
                        "info": {
                            "delegacion": delegacion,
                            "linea_accion": linea_id,
                            "problematica": bloque["problematica"],
                            "lider": bloque["lider"],
                            "rango_inicio": bloque["rango_inicio"],
                            "rango_fin": bloque["rango_fin"],
                        },
                        "tabla": tabla_guardar,
                        "trimestre": trim_value if trim_value in ["", "I", "II", "III", "IV"] else ""
                    }

                st.session_state["pdf_final"] = None
                st.session_state["excel_final"] = None
                st.success("Todas las líneas fueron guardadas correctamente.")

        with gbtn2:
            st.caption("Usa esta opción si deseas guardar de una sola vez todo lo detectado y editado en pantalla.")

        st.markdown("## Líneas guardadas")

        if st.session_state["lineas_guardadas"]:
            cols_saved = st.columns(4)
            ordered_items = list(st.session_state["lineas_guardadas"].items())

            for i, (_, item) in enumerate(ordered_items):
                with cols_saved[i % 4]:
                    st.success(f"Línea {item.get('display_linea', '')}")
        else:
            st.info("Todavía no has guardado ninguna línea.")

        st.markdown("## Resumen consolidado")

        if st.session_state["lineas_guardadas"]:
            df_summary = build_summary_dataframe(st.session_state["lineas_guardadas"])
            st.session_state["ultimo_resumen"] = df_summary.copy()
            st.dataframe(df_summary, use_container_width=True, hide_index=True)
        else:
            st.info("Cuando guardes líneas, aquí verás el resumen consolidado.")

        st.markdown("## Reportes finales")

        p1, p2, p3 = st.columns([2.6, 2.6, 5])

        with p1:
            if st.button("Preparar PDF profesional"):
                if not st.session_state["lineas_guardadas"]:
                    st.warning("Primero debes guardar al menos una línea.")
                else:
                    pdf_bytes = build_pdf_all_lines(
                        st.session_state["lineas_guardadas"],
                        delegacion_general=delegacion,
                        fecha_actualizacion=fecha_actualizacion
                    )
                    st.session_state["pdf_final"] = pdf_bytes
                    st.success("PDF generado correctamente.")

        with p2:
            if st.button("Preparar Excel consolidado"):
                if not st.session_state["lineas_guardadas"]:
                    st.warning("Primero debes guardar al menos una línea.")
                else:
                    excel_bytes = build_excel_export(
                        st.session_state["lineas_guardadas"],
                        delegacion_general=delegacion,
                        fecha_actualizacion=fecha_actualizacion
                    )
                    st.session_state["excel_final"] = excel_bytes
                    st.success("Excel generado correctamente.")

        if st.session_state["pdf_final"]:
            st.download_button(
                "Descargar PDF completo",
                data=st.session_state["pdf_final"],
                file_name=f"reporte_trimestral_{delegacion or 'delegacion'}.pdf",
                mime="application/pdf"
            )

        if st.session_state["excel_final"]:
            st.download_button(
                "Descargar Excel consolidado",
                data=st.session_state["excel_final"],
                file_name=f"reporte_trimestral_{delegacion or 'delegacion'}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with st.expander("Resumen técnico de detección"):
            debug_rows = []
            for b in blocks:
                debug_rows.append({
                    "Línea": b["linea_accion"],
                    "Problemática": b["problematica"],
                    "Líder": b["lider"],
                    "Trimestre detectado": b["trimestre"],
                    "Filas detalle": len(b["tabla"]),
                    "Fila inicio": b["rango_inicio"],
                    "Fila fin": b["rango_fin"],
                    "Fila encabezado": b["header_row"],
                    "Nombre vacío en origen": "Sí" if b.get("nombre_vacio_en_origen") else "No"
                })

            st.dataframe(pd.DataFrame(debug_rows), use_container_width=True, hide_index=True)

    except Exception as e:
        st.error(f"Error al procesar el archivo: {e}")

else:
    st.info("Sube un archivo para comenzar.")
